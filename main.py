import csv
import re
import numpy as np
import matplotlib.pyplot as plt
import pdfkit


from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from enum import Enum
from datetime import datetime
from jinja2 import Environment, FileSystemLoader


"""Enum с полным название валюты"""
class Currency(Enum):
    AZN = "Манаты"
    BYR = "Белорусские рубли"
    EUR = "Евро"
    GEL = "Грузинский лари"
    KGS = "Киргизский сом"
    KZT = "Тенге"
    RUR = "Рубли"
    UAH = "Гривны"
    USD = "Доллары"
    UZS = "Узбекский сум"


"""Enum с текстовым обозначением опыта работы"""
class Experience(Enum):
    noExperience = "Нет опыта"
    between1And3 = "От 1 года до 3 лет"
    between3And6 = "От 3 до 6 лет"
    moreThan6 = "Более 6 лет"


"""Enum с названиями колонок"""
class ColumnName(Enum):
    name = "Название"
    description = "Описание"
    key_skills = "Навыки"
    experience_id = "Опыт работы"
    premium = "Премиум-вакансия"
    employer_name = "Компания"
    salary_from = "Нижняя граница вилки оклада"
    salary_to = "Верхняя граница вилки оклада"
    salary_gross = "Оклад указан до вычета налогов"
    salary_currency = "Идентификатор валюты оклада"
    area_name = "Название региона"
    published_at = "Дата публикации вакансии"


"""Enum с значением конвертации валюты к рублю"""
class CurrencyToRub(Enum):
    AZN = 35.68
    BYR = 23.91
    EUR = 59.90
    GEL = 21.74
    KGS = 0.76
    KZT = 0.13
    RUR = 1
    UAH = 1.64
    USD = 60.66
    UZS = 0.0055


class Salary:
    """Класс для представления зарплат
        Attributes:
            salary_from (int): Нижняя граница оклада
            salary_to (int): Верхняя граница оклада
            salary_currency (int): Валюта оклада
    """
    def __init__(self, salary_from, salary_to, salary_currency):
        """
        :param (str) salary_from:
        :param (str) salary_to:
        :param (str) salary_currency:
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency

    def get_salary_in_rub(self):
        """
        Конвертирует зп в рубли
        :return: (float) Зарплата в рублях
        """
        return ((int(float(self.salary_from)) + int(float(self.salary_to))) / 2) \
               * CurrencyToRub[self.salary_currency].value


class Vacancy:
    """
    Классовое представление вакансии

    Attributes:
        name (str): Имя вакансии
        salary (str): Зарплата
        area_name (str): Название области
        published_at (str): Вребя публикации вакансии
    """
    def __init__(self, name, salary, area_name, published_at):
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class InputConnect:
    """
    Класс для обработки даннх
    """
    @staticmethod
    def connect():
        """
        Метод получения запрашиваемых данных
        :return:
            (int[]) years_salary: годовые зарплата
            (int[]) years_count: количество лет
            (int[]) years_salary_vacancy: зарплаты по выбранной вакансии
            (int[]) years_count_vacancy: количество выбранной вакансии по годам
            (int[]) area_salary: средняя зарплата
            (int[]) area_count: общее количество зп
        """
        file_name = input('Введите название файла: ')
        vac_name = input('Введите название профессии: ')
        years_salary, years_count, years_salary_vacancy, years_count_vacancy, area_salary, area_count \
            = InputConnect.get_data_for_table(DataSet(file_name), vac_name)
        InputConnect\
            .show_data(years_salary, years_count, years_salary_vacancy, years_count_vacancy, area_salary, area_count)
        return years_salary, years_count, years_salary_vacancy, years_count_vacancy, area_salary, area_count, vac_name

    @staticmethod
    def get_data_for_table(dataset, vac_name):
        """
        Метод получения данных из таблицы
        :param dataset: датасет по имени файла
        :param vac_name: название нужнуй вакансии
        :return:
            (int[]) years_salary: годовые зарплата
            (int[]) years_count: количество лет
            (int[]) years_salary_vacancy: зарплаты по выбранной вакансии
            (int[]) years_count_vacancy: количество выбранной вакансии по годам
            (int[]) area_salary: средняя зарплата
            (int[]) area_count: общее количество зп
        """
        vacancies = dataset.vacancies_objects
        years = set()
        for item in vacancies:
            years.add(int(datetime.strptime(item.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime("%Y")))
        years = sorted(list(years))
        years = list(range(min(years), max(years) + 1))

        years_salary = {year: [] for year in years}
        years_count = {year: 0 for year in years}
        years_salary_vacancy = {year: [] for year in years}
        years_count_vacancy = {year: 0 for year in years}

        for item in vacancies:
            formatt = int(datetime.strptime(item.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime("%Y"))
            years_salary[formatt].append(item.salary.get_salary_in_rub())
            years_count[formatt] += 1
            if vac_name in item.name:
                years_salary_vacancy[formatt].append(item.salary.get_salary_in_rub())
                years_count_vacancy[formatt] += 1

        years_salary = {key: int(sum(value) / len(value)) if len(value) != 0 else 0
                        for key, value in years_salary.items()}
        years_salary_vacancy = {key: int(sum(value) / len(value)) if len(value) != 0 else 0
                                for key, value in years_salary_vacancy.items()}

        area = {}
        for item in vacancies:
            if item.area_name in area:
                area[item.area_name].append(item.salary.get_salary_in_rub())
            else:
                area[item.area_name] = [item.salary.get_salary_in_rub()]

        area_list = area.items()
        area_list = [item for item in area_list if len(item[1]) >= int(len(vacancies) / 100)]
        area_list2 = [item for item in area_list if len(item[1]) >= int(len(vacancies) / 100)]
        area_list = sorted(area_list, key=lambda item: sum(item[1]) / len(item[1]), reverse=True)
        area_salary = {item[0]: int(sum(item[1]) / len(item[1])) for item in area_list[0: min(len(area_list), 10)]}

        area_list2 = sorted(area_list2, key=lambda item: len(item[1]) / len(vacancies), reverse=True)
        area_count = {item[0]: round(len(item[1]) / len(vacancies), 4)
                      for item in area_list2[0: min(len(area_list2), 10)]}

        return years_salary, years_count, years_salary_vacancy, years_count_vacancy, area_salary, area_count

    @staticmethod
    def show_data(years_salary: dict, years_count: dict, years_salary_vacancy: dict, years_count_vacancy: dict,
                  area_salary: dict, area_count: dict):
        """
        Метод вывода данных на консоль
        :param (dict) years_salary: годовые зарплата
        :param (dict) years_count: количество лет
        :param (dict) years_salary_vacancy: зарплаты по выбранной вакансии
        :param (dict) years_count_vacancy: количество выбранной вакансии по годам
        :param (dict) area_salary: средняя зарплата
        :param (dict) area_count: общее количество зп
        """
        print(f"Динамика уровня зарплат по годам: {years_salary}")
        print(f"Динамика количества вакансий по годам: {years_count}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {years_salary_vacancy}")
        print(f"Динамика количества вакансий по годам для выбранной профессии: {years_count_vacancy}")
        print(f"Уровень зарплат по городам (в порядке убывания): {area_salary}")
        print(f"Доля вакансий по городам (в порядке убывания): {area_count}")


class Report:
    """
    Класс создания нужного репорта
    """
    @staticmethod
    def generate_excel(years_salary: dict, years_count: dict, years_salary_vacancy: dict,
                       years_count_vacancy: dict, area_salary: dict, area_count: dict, vacancy_name: str):
        """
        Метод создания excel выборки
        :param (dict) years_salary: годовые зарплата
        :param (dict) years_count: количество лет
        :param (dict) years_salary_vacancy: зарплаты по выбранной вакансии
        :param (dict) years_count_vacancy: количество выбранной вакансии по годам
        :param (dict) area_salary: средняя зарплата
        :param (dict) area_count: общее количество зп
        :param (str) vacancy_name: название требуемой вакансии
        """
        wb = Workbook()
        del wb["Sheet"]
        sheet = wb.create_sheet("Статистика по годам")
        side_thin = Side(border_style="thin", color="000000")
        sheet = Report.set_value_in_cell(sheet, "A1", "Год", side_thin)
        sheet = Report.set_value_in_cell(sheet, "B1", "Средняя зарплата", side_thin)
        sheet = Report.set_value_in_cell(sheet, "C1", f"Средняя зарплада - {vacancy_name}", side_thin)
        sheet = Report.set_value_in_cell(sheet, "D1", "Количество вакансий", side_thin)
        sheet = Report.set_value_in_cell(sheet, "E1", f"Количество вакансий - {vacancy_name}", side_thin)
        sheet = Report.set_bold_font(sheet, "A1", "B1", "C1", "D1", "E1")

        for row, (year, value) in enumerate(years_salary.items(), start=2):
            sheet = Report.set_value_in_cell(sheet, f"A{row}", year, side_thin)
            sheet = Report.set_value_in_cell(sheet, f"B{row}", value, side_thin)
            sheet = Report.set_value_in_cell(sheet, f"C{row}", years_salary_vacancy[year], side_thin)
            sheet = Report.set_value_in_cell(sheet, f"D{row}", years_count[year], side_thin)
            sheet = Report.set_value_in_cell(sheet, f"E{row}", years_count_vacancy[year], side_thin)

        for column_cells in sheet.columns:
            length = max(len(Report.try_parse(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        sheet = wb.create_sheet('Статистика по городам')
        sheet = Report.set_value_in_cell(sheet, "A1", "Город", side_thin)
        sheet = Report.set_value_in_cell(sheet, "B1", "Уровень зарплат", side_thin)
        sheet = Report.set_value_in_cell(sheet, "D1", "Город", side_thin)
        sheet = Report.set_value_in_cell(sheet, "E1", "Доля вакансий", side_thin)
        sheet = Report.set_bold_font(sheet, "A1", "B1", "C1", "D1", "E1")

        for row, (year, value) in enumerate(area_salary.items(), start=2):
            sheet = Report.set_value_in_cell(sheet, f"A{row}", year, side_thin)
            sheet = Report.set_value_in_cell(sheet, f"B{row}", value, side_thin)
        for row, (year, value) in enumerate(area_count.items(), start=2):
            sheet = Report.set_value_in_cell(sheet, f"D{row}", year, side_thin)
            sheet = Report.set_value_in_cell(sheet, f"E{row}", value, side_thin)
            sheet[f"E{row}"].number_format = FORMAT_PERCENTAGE_00

        for column_cells in sheet.columns:
            length = max(len(Report.try_parse(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        wb.save('report.xlsx')

    @staticmethod
    def generate_image(years_salary: dict, years_count: dict, years_salary_vacancy: dict,
                       years_count_vacancy: dict, area_salary: dict, area_count: dict, vacancy_name: str):
        """
        Метод генерации изображения по данным
        :param (dict) years_salary: годовые зарплата
        :param (dict) years_count: количество лет
        :param (dict) years_salary_vacancy: зарплаты по выбранной вакансии
        :param (dict) years_count_vacancy: количество выбранной вакансии по годам
        :param (dict) area_salary: средняя зарплата
        :param (dict) area_count: общее количество зп
        :param (str) vacancy_name: название требуемой вакансии
        """
        width_const = 0.4
        _, work_item = plt.subplots(2, 2)
        X_axis = np.arange(len(years_salary.keys()))
        work_item[0, 0] = Report.set_graphic(work_item[0, 0], X_axis, width_const, "средняя з/п", f"з/п {vacancy_name}",
                                             years_salary, years_salary_vacancy, "Уровень зарплат по годам")

        work_item[0, 1] = Report.set_graphic(work_item[0, 1], X_axis, width_const, "Количество вакансий",
                           f"Количество вакансий\n{vacancy_name}", years_count,
                           years_count_vacancy, "Количество вакансий по годам")

        work_item[1, 0].set_title("Уровень зарплат по городам")
        work_item[1, 0].invert_yaxis()
        courses = list(area_salary.keys())
        courses = [label.replace(' ', '\n').replace('-', '-\n') for label in courses]
        work_item[1, 0].tick_params(axis='both', labelsize=8)
        work_item[1, 0].set_yticklabels(courses, fontsize=6, va='center', ha='right')
        work_item[1, 0].barh(courses, list(area_salary.values()))
        work_item[1, 0].grid(True, axis='x')

        other = 1 - sum((list(area_count.values())))
        new_dictionary = {'Другие': other}
        new_dictionary.update(area_count)
        area_count_dic = new_dictionary
        work_item[1, 1].pie(list(area_count_dic.values()), labels=list(area_count_dic.keys()), textprops={'fontsize': 6})
        work_item[1, 1].axis('scaled')
        work_item[1, 1].set_title("Доля вакансий по городам")
        plt.tight_layout()
        plt.savefig('graph.png', dpi=300)

    @staticmethod
    def generate_pdf(years_salary: dict, years_count: dict, years_salary_vacancy: dict,
                       years_count_vacancy: dict, area_salary: dict, area_count: dict, vacancy_name: str):
        """
        Метод генерации pdf по выборке
        :param (dict) years_salary: годовые зарплата
        :param (dict) years_count: количество лет
        :param (dict) years_salary_vacancy: зарплаты по выбранной вакансии
        :param (dict) years_count_vacancy: количество выбранной вакансии по годам
        :param (dict) area_salary: средняя зарплата
        :param (dict) area_count: общее количество зп
        :param (str) vacancy_name: название требуемой вакансии
        """
        area_count = {x[0]: str(f'{x[1] * 100:,.2f}%').replace('.', ',') for x in area_count.items()}

        image_file = "graph.png"
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("index.html")
        header_year = ["Год", "Средняя зарплата", f"Средняя зарплата - {vacancy_name}", "Количество вакансий",
                       f"Количество вакансий - {vacancy_name}"]
        header_city = ["Город", "Уровень зарплат", "Город", "Доля вакансий"]
        pdf_template = template.render({'years_salary': years_salary,
                                        'years_count': years_count,
                                        'years_salary_vacancy': years_salary_vacancy,
                                        'years_count_vacancy': years_count_vacancy, 'area_salary': area_salary,
                                        'area_count': area_count, 'header_year': header_year,
                                        'header_city': header_city, 'image_file': image_file,
                                        'vacancy_name': vacancy_name})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config)

    @staticmethod
    def set_graphic(item, X_axis, width_const, label1, label2, item_bar1, item_bar2, title):
        """
        Метод настройи графики
        :param item: окно
        :param X_axis: отступ по X
        :param width_const: ширина изображения
        :param label1: контент1
        :param label2: контент2
        :param item_bar1: окно контента 1
        :param item_bar2: окно контента 2
        :param title: заголовок
        :return: окно
        """
        item.bar(X_axis - width_const / 2, item_bar1.values(), width=width_const, label=label1)
        item.bar(X_axis + width_const / 2, item_bar2.values(), width=width_const, label=label2)
        item.set_xticks(X_axis, item_bar1.keys())
        item.set_xticklabels(item_bar1.keys(), rotation='vertical', va='top', ha='center')
        item.set_title(title)
        item.grid(True, axis='y')
        item.tick_params(axis='both', labelsize=8)
        item.legend(fontsize=8)

        return item

    @staticmethod
    def set_bold_font(sheet, *cells):
        """
        Метод установки жирного шрифта
        :param sheet: холст
        :param cells: ячейки
        :return: холст
        """
        for cell in cells:
            sheet[cell].font = Font(bold=True)
        return sheet

    @staticmethod
    def set_value_in_cell(sheet, cell: str, value: str, side_thin: Side):
        """
        Метод установки значения в холст
        :param sheet: холст
        :param cell: ячейка
        :param value: значение
        :param side_thin: сторона
        :return:
        """
        sheet[cell] = value
        sheet[cell].border = Border(top=side_thin, left=side_thin, right=side_thin, bottom=side_thin)
        return sheet

    @staticmethod
    def try_parse(item) -> str:
        """
        Метод проверки строки на None
        :param item (str): строка проверки
        :return: сконвертировання строка
        """
        if item is None:
            return ""
        return str(item)


class DataSet:
    """
    Класс структурирования данных

    :argument (str): file_name: Название файла
    :argument (dict): vacancies_objects: Обект с паршенными данными
    """
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = DataSet.parser_csv(file_name)

    @staticmethod
    def clear_str(str_value):
        """
        Метод очищения строки от html
        :param (str): str_value строка для очистки
        :return: очищенная строка
        """
        return ' '.join(re.sub(r"<[^>]+>", '', str_value).split())

    @staticmethod
    def csv_reader(file_name):
        """
        Метод чтения строк из csv
        :param file_name: названия файла
        """
        file = open(file_name, encoding='utf_8_sig')
        reader = [row for row in csv.reader(file)]
        try:
            name = reader.pop(0)
            return name, reader
        except:
            print('Пустой файл')
            exit()

    @staticmethod
    def parser_csv(file_name):
        """
        Метод парсинга строки
        :param (str) file_name: название файла
        :return: распаршенный объект
        """
        naming, reader = DataSet.csv_reader(file_name)
        dict_vacancies = []
        filtered_vacancies = [item for item in reader if len(item) == len(naming) and '' not in item]
        for row in filtered_vacancies:
            dict = {}
            for index in range(0, len(row)):
                if row[index].find("\n") != -1:
                    answer = [DataSet.clear_str(el) for el in row[index].split('\n')]
                else:
                    answer = DataSet.clear_str(row[index])
                dict[naming[index]] = answer
            dict_vacancies.append(
                Vacancy(dict['name'], Salary(dict['salary_from'], dict['salary_to'], dict['salary_currency']),
                        dict['area_name'], dict['published_at']))
        return dict_vacancies


def get_experience_id(value: str):
    """
    Метод получения id от опыта работы
    :param (str) value: опыт работы
    :return: id от опыта работы
    """
    if Experience[value] == Experience.noExperience:
        return 1
    elif Experience[value] == Experience.between1And3:
        return 2
    elif Experience[value] == Experience.between3And6:
        return 3
    elif Experience[value] == Experience.moreThan6:
        return 4
    return 5


true_false = {
    "False": "Нет",
    "True": "Да",
}
true_false_reverse = {
    "Нет": "False",
    "Да": "True",
}

if __name__ == '__main__':
    years_salary, years_count, years_salary_vacancy, years_count_vacancy, area_salary, area_count, vacancy_name \
        = InputConnect.connect()
    choice = input("Вакансии или статистика?").strip().lower()
    if choice == "вакансии":
        Report.generate_excel(years_salary, years_count, years_salary_vacancy, years_count_vacancy, area_salary,
                              area_count, vacancy_name)
        Report.generate_image(years_salary, years_count, years_salary_vacancy, years_count_vacancy, area_salary,
                              area_count, vacancy_name)
        Report.generate_pdf(years_salary, years_count, years_salary_vacancy, years_count_vacancy, area_salary,
                            area_count, vacancy_name)
    else:
        Report.generate_excel(years_salary, years_count, years_salary_vacancy, years_count_vacancy, area_salary,
                              area_count, vacancy_name)
